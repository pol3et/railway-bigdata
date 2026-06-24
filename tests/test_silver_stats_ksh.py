from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from railway_lakehouse.silver.stats import load as stats_load
from railway_lakehouse.silver.stats import merge as stats_merge

pytestmark = pytest.mark.unit


def _xlsx_bytes(path: Path) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "STADAT"

    # Title / note rows imitate real STADAT noise before the data table.
    ws["A1"] = "KSH STADAT railway table"
    ws["A2"] = "Downloaded from the English KSH endpoint"

    # Actual header row: the reader must find years here.
    ws["A4"] = "Indicator"
    ws["B4"] = "Unit"
    ws["C4"] = 2020
    ws["D4"] = 2021

    ws["A5"] = "Rail lines (total route-km)"
    ws["B5"] = "km"
    ws["C5"] = 7891.0
    ws["D5"] = 7869.0

    ws["A6"] = "Rail passengers"
    ws["B6"] = "thousand persons"
    ws["C6"] = 120.0
    ws["D6"] = 150.0

    ws["A7"] = "Road network"
    ws["B7"] = "km"
    ws["C7"] = 31000.0
    ws["D7"] = 32000.0

    wb.save(path)
    return path.read_bytes()


def _wide_year_first_xlsx(path: Path) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "STADAT"

    ws["A1"] = "24.1.1.30. Road and rail network [kilometres]"
    ws["A2"] = "Year"
    ws["B2"] = "National road network total"
    ws["C2"] = "Total length of railway lines operated"
    ws["D2"] = "Length of electrified railway lines"

    ws["A3"] = 2020
    ws["B3"] = 32100
    ws["C3"] = 7670
    ws["D3"] = 3190

    ws["A4"] = 2021
    ws["B4"] = 32200
    ws["C4"] = 7690
    ws["D4"] = 3210

    wb.save(path)
    return path.read_bytes()


def _regional_total_xlsx(path: Path) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "STADAT"

    ws["A1"] = "24.1.2.3. Length of railway lines by county and region [kilometres]"
    ws["A2"] = "Territorial unit denomination"
    ws["B2"] = "Territorial unit level"
    ws["C2"] = 2020
    ws["D2"] = 2021
    ws["A3"] = "Construction length of nationally built standard gauge railways"

    ws["A4"] = "Budapest"
    ws["B4"] = "capital, region"
    ws["C4"] = 194
    ws["D4"] = 194

    ws["A5"] = "Country, total"
    ws["B5"] = "country"
    ws["C5"] = 7606
    ws["D5"] = 7588

    wb.save(path)
    return path.read_bytes()


def _sectioned_single_year_xlsx(path: Path) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "STADAT"

    ws["A1"] = "24.1.1.32. Narrow gauge railway lines"
    ws["A2"] = "Denomination"
    ws["B2"] = 2024
    ws["A3"] = "Means of transport"
    ws["A4"] = "electrified"
    ws["B4"] = 2
    ws["A5"] = "Passenger and goods transport"
    ws["A6"] = "Number of passengers, persons"
    ws["B6"] = 1209478
    ws["A7"] = "Passenger kilometres, thousand"
    ws["B7"] = 15636
    ws["A8"] = "Freight tonne-kilometres, thousand"
    ws["B8"] = 168

    wb.save(path)
    return path.read_bytes()


def test_load_ksh_frame_reads_xlsx_to_long_contract(tmp_path):
    raw = _xlsx_bytes(tmp_path / "sza0030.xlsx")

    frame = stats_load.load_ksh_frame(raw, "ksh_rail_network")

    assert list(frame.columns) == [
        "geo",
        "year",
        "value",
        "unit",
        "source_dataset",
        "source_column",
        "source_system",
    ]
    assert not frame.empty
    assert set(frame["geo"]) == {"HU"}
    assert set(frame["source_system"]) == {"ksh"}
    assert set(frame["unit"]) == {"km", "thousand persons"}
    assert set(frame["source_dataset"]) == {"ksh_rail_network"}

    row = frame[
        (frame["source_column"] == "Rail lines (total route-km)")
        & (frame["year"] == 2021)
    ].iloc[0]
    assert row["value"] == 7869.0


def test_load_ksh_frame_returns_empty_for_non_xlsx_bytes():
    frame = stats_load.load_ksh_frame(b"<html>not an excel file</html>", "broken")
    assert list(frame.columns) == [
        "geo",
        "year",
        "value",
        "unit",
        "source_dataset",
        "source_column",
        "source_system",
    ]
    assert frame.empty


def test_load_ksh_frame_reads_period_year_feature_columns(tmp_path):
    raw = _wide_year_first_xlsx(tmp_path / "sza0030.xlsx")

    frame = stats_load.load_ksh_frame(raw, "ksh_rail_network")

    assert not frame.empty
    assert {
        "National road network total",
        "Total length of railway lines operated",
        "Length of electrified railway lines",
    }.issubset(set(frame["source_column"]))
    assert set(frame["unit"]) == {"kilometres"}

    row = frame[
        (frame["source_column"] == "Total length of railway lines operated")
        & (frame["year"] == 2021)
    ].iloc[0]
    assert row["geo"] == "HU"
    assert row["value"] == 7690


def test_build_silver_stats_maps_ksh_period_year_layout(tmp_path, monkeypatch):
    monkeypatch.setattr(stats_merge, "CROSSWALK_PATH", str(tmp_path / "crosswalk_cache.json"))
    bronze_dir = tmp_path / "stats" / "ksh" / "ksh_rail_network" / "ingest_date=2026-06-23"
    bronze_dir.mkdir(parents=True)
    bronze_dir.joinpath("sza0030.xlsx").write_bytes(_wide_year_first_xlsx(tmp_path / "sza0030.xlsx"))

    unified = stats_load.build_silver_stats(tmp_path, use_llm=False)

    assert set(unified["feature"]) >= {"rail_network_length_km", "rail_electrified_km"}
    assert "National road network total" not in set(unified["source_column"])
    row = unified[
        (unified["feature"] == "rail_network_length_km")
        & (unified["geo"] == "HU")
        & (unified["year"] == 2021)
    ].iloc[0]
    assert row["value"] == 7690
    assert row["unit"] == "kilometres"


def test_load_ksh_frame_uses_country_total_for_regional_tables(tmp_path):
    raw = _regional_total_xlsx(tmp_path / "sza0041.xlsx")

    frame = stats_load.load_ksh_frame(raw, "ksh_rail_lines_regional")

    assert set(frame["source_column"]) == {
        "Construction length of nationally built standard gauge railways"
    }
    assert set(frame["year"]) == {2020, 2021}
    assert set(frame["value"]) == {7606, 7588}
    assert set(frame["unit"]) == {"kilometres"}


def test_load_ksh_frame_keeps_section_context_for_single_year_tables(tmp_path):
    raw = _sectioned_single_year_xlsx(tmp_path / "sza0071.xlsx")

    frame = stats_load.load_ksh_frame(raw, "ksh_rail_narrow_gauge")

    assert "Means of transport - electrified" in set(frame["source_column"])
    assert "electrified" not in set(frame["source_column"])


def test_build_silver_stats_maps_sectioned_ksh_labels_by_metric_not_section(tmp_path, monkeypatch):
    monkeypatch.setattr(stats_merge, "CROSSWALK_PATH", str(tmp_path / "crosswalk_cache.json"))
    bronze_dir = tmp_path / "stats" / "ksh" / "ksh_rail_narrow_gauge" / "ingest_date=2026-06-23"
    bronze_dir.mkdir(parents=True)
    bronze_dir.joinpath("sza0071.xlsx").write_bytes(_sectioned_single_year_xlsx(tmp_path / "sza0071.xlsx"))

    unified = stats_load.build_silver_stats(tmp_path, use_llm=False)

    by_label = dict(zip(unified["source_column"], unified["feature"], strict=False))
    assert by_label["Passenger and goods transport - Number of passengers, persons"] == "rail_passengers"
    assert by_label["Passenger and goods transport - Passenger kilometres, thousand"] == "rail_passenger_km"
    assert by_label["Passenger and goods transport - Freight tonne-kilometres, thousand"] == "rail_freight_tonne_km"
    assert "Means of transport - electrified" not in by_label


def test_frames_from_bronze_picks_up_latest_ksh_partition(tmp_path):
    older = tmp_path / "stats" / "ksh" / "ksh_rail_network" / "ingest_date=2026-06-22"
    newer = tmp_path / "stats" / "ksh" / "ksh_rail_network" / "ingest_date=2026-06-23"
    older.mkdir(parents=True)
    newer.mkdir(parents=True)

    older.joinpath("sza0030.xlsx").write_bytes(_xlsx_bytes(tmp_path / "older.xlsx"))
    newer.joinpath("sza0030.xlsx").write_bytes(_xlsx_bytes(tmp_path / "newer.xlsx"))

    frames = stats_load.frames_from_bronze(tmp_path)

    assert len(frames) == 1
    frame = frames[0]
    assert set(frame["source_system"]) == {"ksh"}
    assert set(frame["source_dataset"]) == {"ksh_rail_network"}


def test_build_silver_stats_maps_known_ksh_label_and_drops_unmapped(tmp_path, monkeypatch):
    monkeypatch.setattr(stats_merge, "CROSSWALK_PATH", str(tmp_path / "crosswalk_cache.json"))

    bronze_dir = tmp_path / "stats" / "ksh" / "ksh_rail_network" / "ingest_date=2026-06-23"
    bronze_dir.mkdir(parents=True)
    bronze_dir.joinpath("sza0030.xlsx").write_bytes(_xlsx_bytes(tmp_path / "sza0030.xlsx"))

    unified = stats_load.build_silver_stats(tmp_path, use_llm=False)

    assert not unified.empty
    assert "rail_network_length_km" in set(unified["feature"])

    row = unified[
        (unified["feature"] == "rail_network_length_km")
        & (unified["geo"] == "HU")
        & (unified["year"] == 2021)
    ].iloc[0]
    assert row["value"] == 7869.0

    cache = pd.read_json(tmp_path / "crosswalk_cache.json", typ="series").to_dict()
    assert cache["Rail lines (total route-km)"] == "rail_network_length_km"

    cache = pd.read_json(tmp_path / "crosswalk_cache.json", typ="series").to_dict()
    assert cache["Rail lines (total route-km)"] == "rail_network_length_km"
